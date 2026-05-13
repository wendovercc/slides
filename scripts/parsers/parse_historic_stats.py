#!/usr/bin/env python3
"""Parse historic batting 100s and bowling 6+ wickets from the club spreadsheet.

Usage:
    python scripts/parsers/parse_historic_stats.py

Input:  data-import/Individual performances to 30-6-25.xlsx
Output: content/data/historic_batting_hundreds.json
        content/data/historic_bowling_sixplus.json
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).parent.parent.parent

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

SPREADSHEET = ROOT / "data-import" / "Individual performances to 30-6-25.xlsx"
OUT_BATTING = ROOT / "content" / "data" / "historic_batting_hundreds.json"
OUT_BOWLING = ROOT / "content" / "data" / "historic_bowling_sixplus.json"

LAST_UPDATED = "2025-06-30"

TEAM_LABELS = {"1": "1st XI", "2": "2nd XI", "3": "3rd XI", "A": "A XI"}


def normalise_team(val):
    return TEAM_LABELS.get(str(val), str(val) if val is not None else None)


def parse_score(val):
    """Return (runs: int, not_out: bool) from a score cell like '170*' or 130."""
    if val is None:
        return None, False
    s = str(val).strip()
    not_out = s.endswith("*")
    num = s.rstrip("*")
    try:
        return int(num), not_out
    except ValueError:
        return None, False


def parse_date(val):
    """Return (iso_date: str|None, year: int|None, approx: bool)."""
    if val is None:
        return None, None, True
    if isinstance(val, datetime):
        # Dates that were entered as just a year (e.g. 01/01/YYYY) have
        # day=1, month=1 — treat those as year-only approximations.
        if val.month == 1 and val.day == 1:
            return None, val.year, True
        return val.date().isoformat(), val.year, False
    s = str(val).strip()
    # "?/1955" or "?/2004" — year known, exact date unknown
    m = re.match(r"^\?/(\d{4})$", s)
    if m:
        return None, int(m.group(1)), True
    return None, None, True


def fmt_overs(val):
    """Format an overs float (e.g. 8.4) as a string, or return None."""
    if val is None:
        return None
    whole = int(val)
    balls = round((val - whole) * 10)
    return f"{whole}.{balls}"


def clean_total(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_batting_hundreds(wb):
    ws = wb["Batting  100's"]
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 2:
            continue
        if all(v is None for v in row):
            continue
        date_val, ha, team_raw, opponents, batsman, score_raw, team_total, oppo_total, result = row[:9]

        # Skip non-data rows (notes, anomaly text)
        if not isinstance(date_val, datetime):
            continue

        runs, not_out = parse_score(score_raw)
        if runs is None:
            continue

        iso_date, year, approx = parse_date(date_val)

        records.append({
            "date": iso_date,
            "year": year,
            "date_approx": approx,
            "home_away": str(ha).strip() if ha is not None else None,
            "team": normalise_team(team_raw),
            "opponents": str(opponents).strip() if opponents else None,
            "batsman": str(batsman).strip() if batsman else None,
            "score": runs,
            "not_out": not_out,
            "team_total": clean_total(team_total),
            "oppo_total": clean_total(oppo_total),
            "result": str(result).strip() if result else None,
        })

    return records


def parse_bowling_sevenpluswkts(wb):
    ws = wb["Bowling - 7+ wkts"]
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 2:
            continue
        if all(v is None for v in row):
            continue
        date_val, ha, team_raw, opponents, bowler, overs_raw, maidens, runs, wickets, wendover, oppo, result = row[:12]

        if not isinstance(date_val, datetime):
            continue
        if wickets is None:
            continue

        iso_date, year, approx = parse_date(date_val)

        records.append({
            "date": iso_date,
            "year": year,
            "date_approx": approx,
            "home_away": str(ha).strip() if ha is not None else None,
            "team": normalise_team(team_raw),
            "opponents": str(opponents).strip() if opponents else None,
            "bowler": str(bowler).strip() if bowler else None,
            "overs": fmt_overs(overs_raw),
            "maidens": int(maidens) if isinstance(maidens, (int, float)) else None,
            "runs": int(runs) if runs is not None else None,
            "wickets": int(wickets),
            "team_total": clean_total(wendover),
            "oppo_total": clean_total(oppo),
            "result": str(result).strip() if result else None,
        })

    return records


def parse_bowling_sixwkts(wb):
    ws = wb["6 wkts"]
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 2:
            continue
        if all(v is None for v in row):
            continue
        date_val, team_raw, opponents, bowler, overs_raw, maidens, runs, wickets = row[:8]

        if date_val is None or wickets is None:
            continue

        iso_date, year, approx = parse_date(date_val)

        # Rows from ~2007 onwards have an H/A marker in the team column ('A' can
        # mean away or A XI — without a dedicated H/A column we can't distinguish,
        # so just normalise the team field as-is)
        records.append({
            "date": iso_date,
            "year": year,
            "date_approx": approx,
            "home_away": None,
            "team": normalise_team(team_raw),
            "opponents": str(opponents).strip() if opponents else None,
            "bowler": str(bowler).strip() if bowler else None,
            "overs": fmt_overs(overs_raw),
            "maidens": int(maidens) if maidens is not None else None,
            "runs": int(runs) if runs is not None else None,
            "wickets": int(wickets),
            "team_total": None,
            "oppo_total": None,
            "result": None,
        })

    return records


def main():
    if not SPREADSHEET.exists():
        sys.exit(f"Spreadsheet not found: {SPREADSHEET}")

    wb = openpyxl.load_workbook(SPREADSHEET)

    # --- Batting 100s ---
    hundreds = parse_batting_hundreds(wb)
    hundreds.sort(key=lambda r: (r["year"] or 0, r["date"] or ""))
    OUT_BATTING.write_text(json.dumps({
        "last_updated": LAST_UPDATED,
        "records": hundreds,
    }, indent=2))
    print(f"Wrote {len(hundreds)} centuries → {OUT_BATTING.relative_to(ROOT)}")

    # --- Bowling 6+ wickets (merge 6-wkt and 7+-wkt sheets) ---
    sixplus = parse_bowling_sixwkts(wb) + parse_bowling_sevenpluswkts(wb)
    sixplus.sort(key=lambda r: (r["year"] or 0, r["date"] or ""))
    OUT_BOWLING.write_text(json.dumps({
        "last_updated": LAST_UPDATED,
        "records": sixplus,
    }, indent=2))
    print(f"Wrote {len(sixplus)} six-plus hauls → {OUT_BOWLING.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
